import datetime
import io
import subprocess

import openpyxl
import pythoncom
from win32com.client import DispatchEx
import platform

from odoo import api, fields, models, _
from calendar import monthrange
from datetime import datetime, time
import time
import math
from dateutil.relativedelta import relativedelta

from odoo.models import MetaModel
from odoo.modules import get_module_path

LIBREOFFICE = 'C:\Program Files\LibreOffice\program\soffice.exe'

import logging

_logger = logging.getLogger(__name__)

class SsSalesInvoiceWizard(models.TransientModel):
    _name = 'ss.sales.invoice.wizard'
    pj_sales_ids = fields.Many2many('ss.sales', string='PJ要員')

    # フォーム初期値
    @api.model
    def default_get(self, field_names):
        defaults = super(SsSalesInvoiceWizard, self).default_get(field_names)
        defaults['pj_sales_ids'] = self.env.context['active_ids']
        return defaults

    # convert to pdf
    def convert2pdf_linux(self, filein, folderout):
        cmd = '"' + LIBREOFFICE + '"' \
              + ' --headless --nologo --nofirststartwizard --norestore  --convert-to pdf --outdir ' \
              + '"' + folderout + '"' + ' ' + '"' + filein + '"'
        p = subprocess.Popen(cmd, stderr=subprocess.PIPE, stdout=subprocess.PIPE)
        p.wait(timeout=60)
        stdout, stderr = p.communicate()
        if stderr:
            raise subprocess.SubprocessError(stderr)
        return stdout

    # MSD請求書-001
    def output_invoice_excel_001(self, pj_id, ws):

        #請求日 'F3'
        ws['H4'].value = datetime.now().strftime('%Y/%m/%d')
        #請求No 'F4'
        ws['H#'].value = pj_id.pj_cd
        #得意先名前 'A5'
        ws['C5'].value = pj_id.pj_partner_c_id.name
        #PJ名 'A16'
        ws['B7'].value = pj_id.pj_name

        #要員詳細
        RowStart = 17
        ws.cell(RowStart,1).value = pj_id.pj_member_id.name

    # MSD請求書-002
    def output_invoice_excel_002(self, pj_id, ws):

        #請求日 'H4'
        ws['H4'].value = pj_id.pj_long_date
        #請求No 'H3'
        ws['H3'].value = pj_id.pj_cd + '-B' + datetime.now().strftime('%m%d')
        #得意先名前 'C5
        ws['C5'].value = pj_id.pj_partner_c_id.name + ' 御中'
        #得意先住所 'C3
        ws['C3'].value = pj_id.pj_partner_c_id.street
        #PJ名 'B7'
        ws['B7'].value = pj_id.pj_name

        #お支払期限 'C12'  基準日＋回収サイクル　例30日
        dt_invoice = datetime.strptime(pj_id.pj_long_date, '%Y-%m-%d')
        cycle=30  #例
        #翌月月末
        if cycle == 30 :
            dt = (dt_invoice.replace(day=1) + relativedelta(months=2, days=-1)).date()
        # 翌翌月5日
        elif cycle == 35 :
            dt = (dt_invoice.replace(day=5) + relativedelta(months=2)).date()
        else :
            dt = (dt_invoice.replace(day=1) + relativedelta(months=1, days=cycle)).date()
        ws['C12'].value = dt.strftime('%Y/%m/%d')

        #要員詳細
        RowStart = 15
        #要員名称
        ws.cell(RowStart,2).value = pj_id.pj_member_id.name
        #数量
        ws.cell(RowStart,4).value = pj_id.pj_o_manhour
        #単位　例：人月
        ws.cell(RowStart,5).value = '人月'
        #単価
        ws.cell(RowStart,6).value = pj_id.pj_o_price_unit
        #摘要　例：作業期間:
        ws.cell(RowStart,8).value = '作業期間:'

        #要員明細2行目
        RowStart = RowStart + 1
        #作業時間　例：作業期間:
        ws.cell(RowStart,3).value = '作業時間:' + str(pj_id.pj_o_duty_hours)
        #要員条件適用
        ws.cell(RowStart,8).value = dt_invoice.strftime('%Y/%m/%d') + '～' + \
            (dt_invoice.replace(day=1) + relativedelta(months=1, days=-1)).date().strftime('%Y/%m/%d')

        #要員条件　控除分/超過分の場合
        if pj_id.pj_o_excess_deduct != 0 :
            #要員明細3行目
            RowStart = RowStart + 1
            ws.cell(RowStart,3).value = '控除分/超過分:' + str(pj_id.pj_o_duty_hours)
            ws.cell(RowStart,6).value = pj_id.pj_o_excess_deduct

        #要員詳細
        # pj_o_price_unit = fields.Integer(u'売価')
        # pj_o_price_purchase = fields.Integer(u'原価')
        # pj_o_manhour_contract = fields.Float(u'標準工数')
        # pj_o_manhour = fields.Float(u'当月工数')
        # pj_o_duty_hours = fields.Float(u'当月時間')
        # pj_o_payoffhour = fields.Float(u'精算時間')
        # pj_o_excess_deduct = fields.Integer(u'超過・控除')
        # pj_o_adjustment = fields.Integer(u'調整額')

        #要員詳細

    def output_sale_pdf_windows(self, record):
        ''''''
        # 変数定義
        file_path = get_module_path('msd_ssm')
        timestamp = str(int(round(time.time() * 1000)))
        file_path_template = file_path + '/static/src/excel/Invoice_template.xlsx'
        file_path_invoice_excel = file_path + '/static/src/excel/Invoice_temp_' + timestamp + '.xlsx'
        file_path_pdf = file_path + '/static/src/excel/Invoice_pdf_' + timestamp + '.pdf'
        file_name_pdf = 'Invoice_pdf_' + timestamp + '.pdf'

        wb = openpyxl.load_workbook(file_path_template)

        # ###############################################
        # Excelファイルへ項目出力
        # ###############################################

        for pj_id in record.pj_sales_ids :
            ws = wb.copy_worksheet(wb.worksheets[0])
            ws.title = pj_id.pj_account_date + '_' + pj_id.pj_cd + '_' + str(pj_id.pj_member_id.name)
            self.output_invoice_excel_002(pj_id, ws)
            # ws.PageSetup.PrintArea = wb.worksheets[0].PageSetup.PrintArea

        # 一時ファイル保存
        wb.save(file_path_invoice_excel)

        # ExcelからPDFに保存する
        pythoncom.CoInitialize()
        xlApp = DispatchEx("Excel.Application")
        xlApp.Visible = False
        xlApp.DisplayAlerts = 0
        books = xlApp.Workbooks.Open(file_path_invoice_excel, False)
        books.ExportAsFixedFormat(0, file_path_pdf)
        books.Close(False)
        xlApp.Quit()

        return {
            'type': 'ir.actions.act_url',
            'name': 'contract',
            'url': '/web/binary/download_document?filename=' + file_name_pdf
        }

    def output_sale_pdf_linux(self, record):
        ''''''
        # 変数定義
        file_path = get_module_path('msd_ssm')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        file_path_template = file_path + '/static/src/excel/Invoice_template.xlsx'
        file_path_temporary = file_path + '/static/src/excel/Invoice_temp_' + timestamp + '.xlsx'
        file_path_pdf = file_path + '/static/src/excel/Invoice_pdf_' + timestamp + '.pdf'
        folder_path_pdf = file_path + '/static/src/excel/'
        file_name_pdf = 'Invoice_temp_' + timestamp + '.pdf'

        wb = openpyxl.load_workbook(file_path_template)
        ws = wb.worksheets[0]

        # ###############################################
        # Excelファイルへ項目出力
        # ###############################################
        self.output_invoice_excel(record, ws)

        # ws.cell(row=5, column=1).value = partner.pj_partner_c_id[0].name
        # # PJコード
        # ws.cell(row=15, column=2).value = partner.pj_cd
        # # 案件名
        # ws.cell(row=16, column=2).value = partner.pj_name

        # 一時ファイル保存
        wb.save(file_path_temporary)

        # ExcelからPDFに保存する
        self.convert2pdf_linux(file_path_temporary, folder_path_pdf)

        return {
            'type': 'ir.actions.act_url',
            'name': 'contract',
            # 'url': '/web/binary/download_document?filename=' + file_name_pdf
            'url': '/msd_ssm/static/src/excel/' + file_name_pdf
        }

    def action_output_sale_pdf(self, record):
        sysstr = platform.system()
        if sysstr =="Windows":
            return self.output_sale_pdf_windows(record);
        elif sysstr == "Linux":
            return self.output_sale_pdf_linux(record);

    def do_create(self):
        self.action_output_sale_pdf(self)
